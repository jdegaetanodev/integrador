import os
import json
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter

RUTA_JSON = os.path.join(os.path.dirname(__file__), "gastosfijos.json")
RUTA_DICT = os.path.join("datos", "gastosfijos.py")
RUTA_EXCEL = os.path.join(os.path.dirname(__file__), "gastosfijos_exportados.xlsx")

def cargar_gastosfijos():
    if not os.path.exists(RUTA_JSON):
        return []
    try:
        with open(RUTA_JSON, "r", encoding="utf-8") as archivo:
            return json.load(archivo)
    except:
        return []

def guardar_gastosfijos(gastosfijos):
    try:
        with open(RUTA_JSON, "w", encoding="utf-8") as archivo:
            json.dump(gastosfijos, archivo, indent=4, ensure_ascii=False)
        guardar_diccionario(gastosfijos)
        return True
    except:
        return False

def guardar_diccionario(gastosfijos):
    try:
        texto = "gastosfijos_dict = {\n"
        for g in gastosfijos:
            texto += f"    {g['id_gasto_fijo']}: {{\n"
            texto += f"        'id_categoria': {g['id_categoria']},\n"
            texto += f"        'nombre': \"{g['nombre']}\",\n"
            texto += f"        'monto': {g['monto']},\n"
            texto += f"        'descripcion': \"{g['descripcion']}\",\n"
            texto += f"        'mes': \"{g['mes']}\",\n"
            texto += f"        'anio': \"{g['anio']}\",\n"
            texto += f"        'abonado': \"{g['abonado']}\",\n"
            texto += f"        'fecha_pago': \"{g['fecha_pago']}\"\n"
            texto += f"    }},\n"
        texto += "}\n"
        with open(RUTA_DICT, "w", encoding="utf-8") as archivo:
            archivo.write(texto)
    except:
        pass

def add_gastofijo(gastofijo):
    try:
        gastosfijos = cargar_gastosfijos()
        nuevo_id = 1
        if gastosfijos:
            nuevo_id = max(g["id_gasto_fijo"] for g in gastosfijos) + 1
        gastofijo["id_gasto_fijo"] = nuevo_id
        gastosfijos.append(gastofijo)
        return guardar_gastosfijos(gastosfijos)
    except:
        return False

def update_gastofijo(id_gastofijo, nuevo_gastofijo):
    try:
        gastosfijos = cargar_gastosfijos()
        for g in gastosfijos:
            if g["id_gasto_fijo"] == id_gastofijo:
                g.update(nuevo_gastofijo)
                return guardar_gastosfijos(gastosfijos)
        return False
    except:
        return False

def delete_gastofijo(id_gastofijo):
    try:
        gastosfijos = cargar_gastosfijos()
        nuevos = [g for g in gastosfijos if g["id_gasto_fijo"] != id_gastofijo]
        if len(nuevos) == len(gastosfijos):
            return False
        return guardar_gastosfijos(nuevos)
    except:
        return False
    
def exportar_gastos_fijos():
    try:
        # Cargar los datos de los gastos fijos
        gastosfijos = cargar_gastosfijos()
        if not gastosfijos:
            print("No hay gastos fijos para exportar.")
            return False

        # Crear el libro y la hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos Fijos"

        # Definir los encabezados
        encabezados = ["Oden", "Categoría", "Nombre", "Monto", "Descripción", "Mes", "Año", "Abonado", "Fecha de Pago"]
        ws.append(encabezados)

        # Rellenar los datos
        for gasto in gastosfijos:
            fila = [
                gasto['id_gasto_fijo'],
                gasto['id_categoria'],
                gasto['nombre'],
                gasto['monto'],
                gasto['descripcion'],
                gasto['mes'],
                gasto['anio'],
                "Sí" if gasto['abonado'] else "No",
                gasto['fecha_pago']
            ]
            ws.append(fila)

        # Ajustar el ancho de las columnas automáticamente
        for col in range(1, len(encabezados) + 1):
            columna = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[columna].width = adjusted_width

        # Definir la ruta del archivo exportado
        ruta_archivo = os.path.join(os.path.dirname(__file__), "gastosfijos_exportados.xlsx")
        
        # Guardar el archivo
        wb.save(ruta_archivo)
        
        # Mensaje de éxito con la ruta del archivo
        messagebox.showinfo("Éxito", f"Archivo exportado con éxito a:\n{ruta_archivo}")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar gastos fijos a Excel:\n{e}")


def validar_datos():
    return True         