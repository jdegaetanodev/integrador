import os
import json
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter

RUTA_JSON = os.path.join(os.path.dirname(__file__), "presupuesto.json")
RUTA_DICT = os.path.join("datos", "presupuesto.py")
RUTA_EXCEL = os.path.join(os.path.dirname(__file__), "presupuesto_exportados.xlsx")

def cargar_presupuestos():
    if not os.path.exists(RUTA_JSON):
        return []
    try:
        with open(RUTA_JSON, "r", encoding="utf-8") as archivo:
            presupuestos = json.load(archivo)
        return presupuestos
    except:
        return []

def guardar_presupuestos(presupuestos):
    try:
        with open(RUTA_JSON, "w", encoding="utf-8") as archivo:
            json.dump(presupuestos, archivo, indent=4, ensure_ascii=False)
        guardar_diccionario(presupuestos)
        return True
    except:
        return False

def guardar_diccionario(presupuestos):
    try:
        texto = "presupuesto_dict = {\n"
        for p in presupuestos:
            texto += f"    {p['id_presupuesto']}: {{\n"
            texto += f"        'id_categoria': {p['id_categoria']},\n"
            texto += f"        'monto_presupuesto': {p['monto_presupuesto']},\n"
            texto += f"        'descripcion': \"{p['descripcion']}\",\n"
            texto += f"        'mes': {p['mes']},\n"
            texto += f"        'anio': {p['anio']}\n"
            texto += f"    }},\n"
        texto += "}\n"
        with open(RUTA_DICT, "w", encoding="utf-8") as archivo:
            archivo.write(texto)
    except:
        pass  

def add_presupuesto(presupuesto):
    """
    presupuesto es un dict con keys: id_categoria, monto_presupuesto, descripcion, mes, anio
    """
    try:
        presupuestos = cargar_presupuestos()
        nuevo_id = 1
        if presupuestos:
            nuevo_id = max(p["id_presupuesto"] for p in presupuestos) + 1
        presupuesto["id_presupuesto"] = nuevo_id
        presupuestos.append(presupuesto)
        return guardar_presupuestos(presupuestos)
    except:
        return False

def update_presupuesto(id_presupuesto, nuevo_presupuesto):
    """
    nuevo_presupuesto es un dict con keys: id_categoria, monto_presupuesto, descripcion, mes, anio
    """
    try:
        presupuestos = cargar_presupuestos()
        for p in presupuestos:
            if p["id_presupuesto"] == id_presupuesto:
                p.update(nuevo_presupuesto)
                return guardar_presupuestos(presupuestos)
        return False
    except:
        return False

def delete_presupuesto(id_presupuesto):
    try:
        presupuestos = cargar_presupuestos()
        nuevos_presupuestos = [p for p in presupuestos if p["id_presupuesto"] != id_presupuesto]
        if len(nuevos_presupuestos) == len(presupuestos):
            return False
        return guardar_presupuestos(nuevos_presupuestos)
    except:
        return False

def exportar_presupuesto():
    try:
        # Cargar los datos de los presupuestos
        presupuestos = cargar_presupuestos()
        if not presupuestos:
            print("No hay presupuestos para exportar.")
            return False

        # Crear el libro y la hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presupuestos"

        # Definir los encabezados
        encabezados = ["Orden", "Categoría", "Monto Presupuesto", "Descripción", "Mes", "Año"]
        ws.append(encabezados)

        # Rellenar los datos
        for presupuesto in presupuestos:
            fila = [
                presupuesto['id_presupuesto'],
                presupuesto['id_categoria'],
                presupuesto['monto_presupuesto'],
                presupuesto['descripcion'],
                presupuesto['mes'],
                presupuesto['anio']
            ]
            ws.append(fila)

        # Ajustar el ancho de las columnas automáticamente
        for col in range(1, len(encabezados) + 1):
            columna = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[columna].width = adjusted_width

        # Definir la ruta del archivo exportado
        ruta_archivo = os.path.join(os.path.dirname(__file__), "presupuesto_exportados.xlsx")
        
        # Guardar el archivo
        wb.save(ruta_archivo)
        
        # Mensaje de éxito con la ruta del archivo
        messagebox.showinfo("Éxito", f"Archivo exportado con éxito a:\n{ruta_archivo}")
        return True  # Aseguramos que la función retorne True al final
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar presupuestos a Excel:\n{e}")
        return False