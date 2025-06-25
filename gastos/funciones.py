import os
import json
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter

RUTA_JSON = os.path.join(os.path.dirname(__file__), "gastos.json")
RUTA_DICT = os.path.join("datos", "gastos.py")
RUTA_EXCEL = os.path.join(os.path.dirname(__file__), "gastos_exportados.xlsx")


def cargar_gastos():
    if not os.path.exists(RUTA_JSON):
        return []
    try:
        with open(RUTA_JSON, "r", encoding="utf-8") as archivo:
            gastos = json.load(archivo)
        return gastos
    except:
        return []

def guardar_gastos(gastos):
    try:
        # Crear carpeta si no existe
        os.makedirs(os.path.dirname(RUTA_JSON), exist_ok=True)

        with open(RUTA_JSON, "w", encoding="utf-8") as archivo:
            json.dump(gastos, archivo, indent=4, ensure_ascii=False)
            archivo.flush()
            os.fsync(archivo.fileno())

        return True
    except Exception as e:
        print("Error en guardar_gastos:", e)
        return False

def guardar_diccionario(gastos):
    try:
        texto = "gastos = [\n"
        for g in gastos:
            texto += "    {\n"
            texto += f"        'id_gastos': {g['id_gastos']},\n"
            texto += f"        'id_categoria': {g['id_categoria']},\n"
            texto += f"        'nombre': \"{g['nombre']}\",\n"
            texto += f"        'monto': {g['monto']},\n"
            texto += f"        'detalle': \"{g['detalle']}\",\n"
            texto += f"        'fecha': \"{g['fecha']}\"\n"
            texto += "    },\n"
        texto += "]\n"
        with open(RUTA_DICT, "w", encoding="utf-8") as archivo:
            archivo.write(texto)
    except:
        pass

def add_gasto(nombre, id_categoria, monto, fecha, descripcion, ventana_gastos, callback_actualizar=None):   

    if validar_datos(id_categoria, nombre, monto, fecha, descripcion):   

        try:
            gastos = cargar_gastos()
            nuevo_id = 1
            if gastos:
                nuevo_id = max(c["id_gastos"] for c in gastos) + 1

            nuevo_gasto = {
                "id_gastos": nuevo_id,
                "nombre": nombre,
                "id_categoria": id_categoria,
                "monto": monto,
                "fecha": fecha,
                "descripcion": descripcion
            }

            gastos.append(nuevo_gasto)
            guardado = guardar_gastos(gastos)

            if guardado:
                messagebox.showinfo("Éxito", "Gasto cargado correctamente.")
                if callback_actualizar:
                    callback_actualizar()
                ventana_gastos.destroy()
                return True
            else:
                messagebox.showerror("Error", "No se pudo guardar el gasto.")
                return False
        except Exception as e:
            print("Error en add_gasto:", e)
            messagebox.showerror("Error", "Ocurrió un error al añadir el gasto.")
            return False
    else:
        messagebox.showerror("Error", "Por favor complete todos los campos del formulario")


def update_gasto(id_gasto, id_categoria, nombre, categoria, monto, fecha, detalle, ventana_gasto, callback_actualizar):

    if validar_datos(id_categoria, nombre, monto, fecha, detalle):   

        try:
            nuevo_gasto = {
                "id_categoria": int(id_categoria),
                "nombre": nombre,
                "categoria": categoria,
                "monto": monto,
                "fecha": fecha,
                "detalle": detalle
            }

            gastos = cargar_gastos()
            for gasto in gastos:

                if gasto["id_gastos"] == int(id_gasto):
                    gasto.update(nuevo_gasto)
                    guardado = guardar_gastos(gastos)
                    if guardado:
                        messagebox.showinfo("Éxito", "Gasto actualizado correctamente.")
                        if callback_actualizar:
                            callback_actualizar()
                        ventana_gasto.destroy()
                        return True
                    else:
                        messagebox.showerror("Error", "No se pudo guardar el gasto.")
                        return False
            messagebox.showerror("Error", "No se encontró un Gasto con ese ID.")
            return False
        except Exception as e:
            print("Error en update_gasto:", e)
            messagebox.showerror("Error", "Ocurrió un error al actualizar el gasto.")
            return False
    else:
        messagebox.showerror("Error", "Por favor complete todos los campos del formulario")

def delete_gasto(id_gasto):
    try:
        gastos = cargar_gastos()
        nuevos_gastos = [g for g in gastos if g["id_gastos"] != id_gasto]
        if len(nuevos_gastos) == len(gastos):
            return False
        return guardar_gastos(nuevos_gastos)
    except:
        return False

def exportar_gastos():
    try:
        # Cargar los datos de los gastos
        gastos = cargar_gastos()
        if not gastos:
            print("No hay gastos para exportar.")
            return False

        # Crear el libro y la hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos"

        # Definir los encabezados
        encabezados = ["Orden", "Categoría", "Nombre", "Monto", "Detalle", "Fecha"]
        ws.append(encabezados)

        # Rellenar los datos
        for gasto in gastos:
            fila = [
                gasto['id_gastos'],
                gasto['id_categoria'],
                gasto['nombre'],
                gasto['monto'],
                gasto['detalle'],
                gasto['fecha']
            ]
            ws.append(fila)

        # Ajustar el ancho de las columnas automáticamente
        for col in range(1, len(encabezados) + 1):
            columna = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[columna].width = adjusted_width

           # Definir la ruta del archivo exportado
        ruta_archivo = os.path.join(os.path.dirname(__file__), "gastos_exportados.xlsx")
        
        # Guardar el archivo
        wb.save(ruta_archivo)
        
        # Mensaje de éxito con la ruta del archivo
        messagebox.showinfo("Éxito", f"Archivo exportado con éxito a:\n{ruta_archivo}")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar gastos a Excel:\n{e}")

def validar_datos(id_categoria, nombre, monto, fecha, detalle):    
    return True      